from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.models import RunStatus, ShippingItem, ShippingOrder, ShippingState
from app.repository import ShippingJobRepository
from app.shipping.orchestrator import ShippingOrchestrator
from app.shipping.excel_generator import WmsExcelGenerator
from app.shipping.errors import RetryableAutomationError
from app.shipping.pdf_handler import extract_tracking_number, find_tracking_number, normalize_label
from app.shipping.sku_mapper import SkuMapper


def make_settings(tmp_path: Path, mapping_text: str) -> Settings:
    mapping_path = tmp_path / "product_mapping.csv"
    mapping_path.write_text(mapping_text, encoding="utf-8")
    return Settings(
        base_dir=tmp_path,
        data_dir=tmp_path,
        database_path=tmp_path / "jobs.sqlite3",
        artifacts_dir=tmp_path / "artifacts",
        sku_mapping_path=mapping_path,
        wms_template_path=None,
        automation_mode="dry-run",
        headless=True,
        tiktok_orders_url="",
        tiktok_profile_dir=tmp_path / "profiles" / "tiktok",
        wms_login_url="",
        wms_upload_url="",
        wms_orders_url="",
        wms_label_upload_url="",
        wms_username="",
        wms_password="",
        wms_profile_dir=tmp_path / "profiles" / "wms",
        dry_run_order_id="TKE-TEST-SWEEP",
    )


def make_order(
    order_id: str = "TKE-TEST-001",
    sku: str = "XCGLM-GLM851",
    product_title: str = "Estrella Hair Kinky Curly Bundles",
    variant_name: str = "14A-Kinky curly-18",
) -> ShippingOrder:
    return ShippingOrder(
        platform_order_id=order_id,
        country="US",
        transport_method="CBT-DF",
        items=[
            ShippingItem(
                platform_sku=sku,
                product_title=product_title,
                variant_name=variant_name,
                quantity=1,
            )
        ],
    )


def test_dry_run_completes_entire_shipping_pipeline(tmp_path: Path) -> None:
    settings = make_settings(
        tmp_path,
        "平台SKU,产品编号,中文名称\nXCGLM-GLM851,GLM851,14A-Kinky curly-18\n",
    )
    repository = ShippingJobRepository(settings.database_path)
    orchestrator = ShippingOrchestrator(settings, repository)

    created = orchestrator.create_job(make_order())
    completed = orchestrator.run_job(created.id)

    assert completed.state == ShippingState.COMPLETED
    assert completed.run_status == RunStatus.COMPLETED
    assert completed.tracking_number.startswith("SWX")
    assert Path(completed.label_path).name == f"{completed.tracking_number}.pdf"
    assert Path(completed.label_path).exists()
    assert Path(completed.excel_path).exists()
    assert completed.mapped_items[0]["product_code"] == "GLM851"
    assert completed.mapped_items[0]["fulfillment_sku"] == "XCGLM-GLM851"
    assert completed.mapped_items[0]["matched_by"] == "platform_sku"
    assert completed.events[-1]["state"] == ShippingState.COMPLETED


def test_completed_job_is_idempotent(tmp_path: Path) -> None:
    settings = make_settings(
        tmp_path,
        "平台SKU,产品编号,中文名称\nXCGLM-GLM851,GLM851,14A-Kinky curly-18\n",
    )
    repository = ShippingJobRepository(settings.database_path)
    orchestrator = ShippingOrchestrator(settings, repository)
    job = orchestrator.create_job(make_order())

    first = orchestrator.run_job(job.id)
    event_count = len(first.events)
    second = orchestrator.run_job(job.id)

    assert second.state == ShippingState.COMPLETED
    assert len(second.events) == event_count


def test_more_than_five_products_force_hc_us_transport_method() -> None:
    order = ShippingOrder(
        platform_order_id="TKE-BULK-001",
        transport_method="CBT-DF",
        items=[
            ShippingItem(product_title="Product A", quantity=3),
            ShippingItem(product_title="Product B", quantity=3),
        ],
    )

    assert order.total_quantity() == 6
    assert order.transport_method == "HC-US"


def test_exactly_five_products_preserve_transport_method() -> None:
    order = ShippingOrder(
        platform_order_id="TKE-STANDARD-001",
        transport_method="CBT-DF",
        items=[
            ShippingItem(product_title="Product A", quantity=3),
            ShippingItem(product_title="Product B", quantity=2),
        ],
    )

    assert order.total_quantity() == 5
    assert order.transport_method == "CBT-DF"


def test_unknown_sku_stops_for_manual_intervention(tmp_path: Path) -> None:
    settings = make_settings(
        tmp_path,
        "平台SKU,产品编号,中文名称\nXCGLM-GLM851,GLM851,14A-Kinky curly-18\n",
    )
    repository = ShippingJobRepository(settings.database_path)
    orchestrator = ShippingOrchestrator(settings, repository)
    job = orchestrator.create_job(
        make_order(sku="UNKNOWN-SKU", product_title="Unknown product", variant_name="Unknown variant")
    )

    result = orchestrator.run_job(job.id)

    assert result.state == ShippingState.LABEL_DOWNLOADED
    assert result.run_status == RunStatus.MANUAL_INTERVENTION_REQUIRED
    assert result.error_code == "sku_mapping_error"
    assert "UNKNOWN-SKU" in result.error_message


def test_sweep_discovers_and_completes_order(tmp_path: Path) -> None:
    settings = make_settings(
        tmp_path,
        "平台SKU,产品编号,中文名称\n"
        "XCGLM-GLM851,GLM851,14A-Kinky curly-18\n"
        "BONUS-LASH,zeng1,Limited Free Bonus Eyelash Clusters\n",
    )
    repository = ShippingJobRepository(settings.database_path)
    orchestrator = ShippingOrchestrator(settings, repository)

    jobs = orchestrator.sweep()

    assert len(jobs) == 1
    assert jobs[0].platform_order_id == "TKE-TEST-SWEEP"
    assert jobs[0].run_status == RunStatus.COMPLETED
    assert [item["product_code"] for item in jobs[0].mapped_items] == ["GLM851", "zeng1"]


def test_mapper_uses_product_total_sheet_and_real_headers(tmp_path: Path) -> None:
    path = tmp_path / "product.xlsx"
    workbook = Workbook()
    workbook.active.title = "分类1"
    workbook.active.append(["产品SKU(SKU)", "产品条码(product barcode)", "中文名称(product name in Chinese)"])
    workbook.active.append(["WRONG-SKU", "WRONG", "Wrong product"])
    sheet = workbook.create_sheet("产品总")
    sheet.append(
        [
            "品类ID(choose category)",
            "产品SKU(SKU)",
            "产品条码(product barcode)",
            "中文名称(product name in Chinese)",
            "英文申报品名(item name in English)",
            "中文申报品名(item name in Chinese)",
            "海关编码(HS code)",
            "USD申报价值(USD declared value)",
            "EUR申报价值(EUR declared value)",
            "产品单重(KG)(weight)",
        ]
    )
    sheet.append([70806, "XCGLM-GLM851", "GLM851", "14A-Kinky curly-18", "hair", "头套", 6704200000, 99, "", 0.2])
    workbook.save(path)

    mapped = SkuMapper(path).map_items(
        [ShippingItem(platform_sku="GLM851", product_title="Kinky Curly")]
    )

    assert mapped[0]["matched_by"] == "product_barcode"
    assert mapped[0]["fulfillment_sku"] == "XCGLM-GLM851"
    assert mapped[0]["product_code"] == "GLM851"
    assert mapped[0]["product_name"] == "14A-Kinky curly-18"


def test_real_wms_columns_and_existing_rows_are_replaced(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "客户单号", "转单号", "运输方式", "国家", "收件公司", "收件人", "州省", "城市", "地址", "电话", "邮编", "重量",
        "中文名1", "英文名1", "配货1", "数量1", "申报1",
    ]
    sheet.append(headers)
    sheet.append(["OLD-ORDER", "OLD-TRACKING"])
    workbook.save(template_path)
    generator = WmsExcelGenerator(tmp_path / "out", template_path, item_slots=2)
    order = make_order()
    mapped_items = [
        {
            "fulfillment_sku": "XCGLM-GLM851",
            "product_code": "GLM851",
            "quantity": 2,
            "declaration_quantity": 2,
        }
    ]

    result = generator.generate(order, "9200190417700000000001", mapped_items)
    result_sheet = load_workbook(result, data_only=True).active
    values = {result_sheet.cell(1, col).value: result_sheet.cell(2, col).value for col in range(1, result_sheet.max_column + 1)}

    assert values["客户单号"] == order.platform_order_id
    assert values["转单号"] == "9200190417700000000001"
    assert values["收件人"] == "1"
    assert values["重量"] == 0.1
    assert values["中文名1"] == "头套"
    assert values["英文名1"] == "wig"
    assert values["配货1"] == "XCGLM-GLM851"
    assert values["数量1"] == 2
    assert values["申报1"] == 2
    assert "配货2" in values


def test_tracking_number_supports_usps_and_platform_formats() -> None:
    assert find_tracking_number("9200 1904 1770 0000 0000 01") == "9200190417700000000001"
    assert find_tracking_number("Tracking SWX148730000106223600") == "SWX148730000106223600"


def test_image_only_label_can_use_tracking_filename(tmp_path: Path) -> None:
    label = tmp_path / "9200190417700000000001.pdf"
    label.write_bytes(b"image-only-pdf-placeholder")

    assert extract_tracking_number(label) == "9200190417700000000001"
    normalized = normalize_label(label, "9200190417700000000001", tmp_path / "labels")
    assert normalized.exists()
    try:
        normalize_label(label, "SWX148730000106223600", tmp_path / "wrong")
    except RetryableAutomationError as error:
        assert "mismatch" in str(error)
    else:
        raise AssertionError("tracking mismatch must stop the workflow")
