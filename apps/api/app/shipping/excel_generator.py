from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from ..models import ShippingOrder
from .errors import ConfigurationError


BASE_HEADERS = [
    "客户单号",
    "转单号",
    "运输方式",
    "国家",
    "收件公司",
    "收件人",
    "州省",
    "城市",
    "地址",
    "电话",
    "邮编",
    "重量",
]

ITEM_HEADER_NAMES = ("中文名", "英文名", "配货", "数量", "申报")


class WmsExcelGenerator:
    def __init__(
        self,
        output_dir: Path,
        template_path: Path | None = None,
        *,
        placeholder_value: str = "1",
        default_weight: float = 0.1,
        declaration_cn: str = "头套",
        declaration_en: str = "wig",
        item_slots: int = 10,
    ):
        self.output_dir = output_dir
        self.template_path = template_path
        self.placeholder_value = placeholder_value
        self.default_weight = default_weight
        self.declaration_cn = declaration_cn
        self.declaration_en = declaration_en
        self.item_slots = item_slots
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        order: ShippingOrder,
        tracking_number: str,
        mapped_items: list[dict[str, Any]],
    ) -> Path:
        if not mapped_items:
            raise ConfigurationError("Cannot generate WMS workbook without mapped items")

        if self.template_path:
            if not self.template_path.exists():
                raise ConfigurationError(f"WMS template does not exist: {self.template_path}")
            if self.template_path.suffix.lower() != ".xlsx":
                raise ConfigurationError("The initial WMS template adapter requires an .xlsx template")
            workbook = load_workbook(self.template_path)
            sheet = workbook.active
            header_row = self._find_header_row(sheet)
            self._ensure_item_slots(sheet, header_row, max(self.item_slots, len(mapped_items)))
            self._clear_existing_rows(sheet, header_row)
            self._fill_existing_template(sheet, header_row, order, tracking_number, mapped_items)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "订单导入"
            dynamic_headers = self._item_headers(max(self.item_slots, len(mapped_items)))
            sheet.append(BASE_HEADERS + dynamic_headers)
            self._fill_default_sheet(sheet, 2, order, tracking_number, mapped_items)

        path = self.output_dir / f"hcrd_{order.normalized_customer_order_no()}_{tracking_number}.xlsx"
        workbook.save(path)
        return path

    @staticmethod
    def _find_header_row(sheet) -> int:
        for row_index in range(1, min(sheet.max_row, 20) + 1):
            values = {str(sheet.cell(row_index, column).value or "").strip() for column in range(1, sheet.max_column + 1)}
            if "客户单号" in values and ("转单号" in values or "物流号" in values):
                return row_index
        raise ConfigurationError("Could not locate WMS template header row")

    def _fill_existing_template(
        self, sheet, header_row: int, order: ShippingOrder, tracking: str, items: list[dict[str, Any]]
    ) -> None:
        headers = {
            str(sheet.cell(header_row, column).value or "").strip(): column
            for column in range(1, sheet.max_column + 1)
        }
        row = header_row + 1
        values = self._base_values(order, tracking)
        for header, value in values.items():
            column = headers.get(header)
            if column:
                sheet.cell(row, column, value)
        for index, item in enumerate(items, 1):
            aliases = {
                f"中文名{index}": self.declaration_cn,
                f"英文名{index}": self.declaration_en,
                f"配货{index}": item["fulfillment_sku"],
                f"数量{index}": item["quantity"],
                f"申报{index}": item["declaration_quantity"],
            }
            for header, value in aliases.items():
                column = headers.get(header)
                if column:
                    sheet.cell(row, column, value)

    def _fill_default_sheet(
        self, sheet, row: int, order: ShippingOrder, tracking: str, items: list[dict[str, Any]]
    ) -> None:
        base = list(self._base_values(order, tracking).values())
        item_values: list[Any] = []
        for item in items:
            item_values.extend(
                [
                    self.declaration_cn,
                    self.declaration_en,
                    item["fulfillment_sku"],
                    item["quantity"],
                    item["declaration_quantity"],
                ]
            )
        for column, value in enumerate(base + item_values, 1):
            sheet.cell(row, column, value)

    def _base_values(self, order: ShippingOrder, tracking: str) -> dict[str, Any]:
        recipient = order.recipient
        placeholder = self.placeholder_value
        return {
            "客户单号": order.normalized_customer_order_no(),
            "转单号": tracking,
            "运输方式": order.transport_method,
            "国家": order.country,
            "收件公司": "",
            "收件人": recipient.name or placeholder,
            "州省": recipient.state or placeholder,
            "城市": recipient.city or placeholder,
            "地址": recipient.address_line_1 or placeholder,
            "电话": recipient.phone or placeholder,
            "邮编": recipient.postal_code or placeholder,
            "重量": self.default_weight,
        }

    @staticmethod
    def _item_headers(count: int) -> list[str]:
        return [f"{name}{index}" for index in range(1, count + 1) for name in ITEM_HEADER_NAMES]

    @staticmethod
    def _clear_existing_rows(sheet, header_row: int) -> None:
        for row in sheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.value = None

    @staticmethod
    def _ensure_item_slots(sheet, header_row: int, count: int) -> None:
        headers = {
            str(sheet.cell(header_row, column).value or "").strip(): column
            for column in range(1, sheet.max_column + 1)
        }
        for index in range(1, count + 1):
            for field_offset, name in enumerate(ITEM_HEADER_NAMES):
                header = f"{name}{index}"
                if header not in headers:
                    column = len(BASE_HEADERS) + (index - 1) * len(ITEM_HEADER_NAMES) + field_offset + 1
                    source_index = max(1, min(index - 1, 5))
                    source_header = f"{name}{source_index}"
                    source_column = headers.get(source_header)
                    if source_column:
                        for row_index in range(1, max(sheet.max_row, header_row + 1) + 1):
                            source = sheet.cell(row_index, source_column)
                            target = sheet.cell(row_index, column)
                            if source.has_style:
                                target._style = copy(source._style)
                            target.number_format = source.number_format
                            target.alignment = copy(source.alignment)
                        source_letter = get_column_letter(source_column)
                        target_letter = get_column_letter(column)
                        sheet.column_dimensions[target_letter].width = sheet.column_dimensions[source_letter].width
                    sheet.cell(header_row, column).value = header
                    headers[header] = column
