from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from ..models import ShippingItem
from .errors import ConfigurationError, SkuMappingError


PLATFORM_SKU_HEADERS = {"平台sku", "sku", "卖家sku", "seller sku", "platform sku"}
PRODUCT_CODE_HEADERS = {"产品编号", "product", "product code", "内部产品编号", "货品编号"}
PRODUCT_NAME_HEADERS = {"中文名称", "产品名称", "product name", "名称", "商品名称"}
DECLARATION_EN_HEADERS = {"英文申报品名", "item name in english"}
DECLARATION_CN_HEADERS = {"中文申报品名", "item name in chinese"}
DECLARED_VALUE_HEADERS = {"usd申报价值", "usd declared value"}
WEIGHT_HEADERS = {"产品单重", "weight"}


def _normalize(value: Any) -> str:
    return re.sub(r"[\s_\-/]+", "", str(value or "").strip().lower())


def _find_index(headers: list[str], candidates: set[str]) -> int | None:
    normalized_candidates = {_normalize(candidate) for candidate in candidates}
    for index, header in enumerate(headers):
        normalized_header = _normalize(header)
        if any(
            normalized_header == candidate
            or normalized_header.startswith(candidate + "(")
            or candidate in normalized_header
            for candidate in normalized_candidates
        ):
            return index
    return None


class SkuMapper:
    def __init__(self, mapping_path: Path):
        self.mapping_path = mapping_path
        self._by_sku: dict[str, dict[str, str]] = {}
        self._by_barcode: dict[str, dict[str, str]] = {}
        self._by_name: dict[str, dict[str, str]] = {}
        self.reload()

    def reload(self) -> None:
        if not self.mapping_path.exists():
            raise ConfigurationError(f"SKU mapping file does not exist: {self.mapping_path}")

        suffix = self.mapping_path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            delimiter = "\t" if suffix == ".tsv" else ","
            with self.mapping_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle, delimiter=delimiter))
        elif suffix == ".xlsx":
            workbook = load_workbook(self.mapping_path, read_only=True, data_only=True)
            worksheet = workbook["产品总"] if "产品总" in workbook.sheetnames else workbook.active
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        elif suffix == ".xls":
            try:
                import xlrd
            except ImportError as exc:  # pragma: no cover - dependency is declared
                raise ConfigurationError("xlrd is required to read .xls mapping files") from exc
            workbook = xlrd.open_workbook(self.mapping_path)
            sheet = (
                workbook.sheet_by_name("产品总")
                if "产品总" in workbook.sheet_names()
                else workbook.sheet_by_index(0)
            )
            rows = [sheet.row_values(index) for index in range(sheet.nrows)]
        else:
            raise ConfigurationError(f"Unsupported SKU mapping format: {suffix}")

        if not rows:
            raise ConfigurationError("SKU mapping file is empty")

        header_row_index, header_indexes = self._locate_headers(rows)
        self._by_sku.clear()
        self._by_barcode.clear()
        self._by_name.clear()
        for row in rows[header_row_index + 1 :]:
            record = self._record(row, header_indexes)
            if not record["product_code"]:
                continue
            if record["platform_sku"]:
                self._by_sku[_normalize(record["platform_sku"])] = record
            if record["product_code"]:
                self._by_barcode[_normalize(record["product_code"])] = record
            if record["product_name"]:
                self._by_name[_normalize(record["product_name"])] = record

        if not self._by_sku and not self._by_name:
            raise ConfigurationError("SKU mapping file contains no usable mapping rows")

    def map_items(self, items: Iterable[ShippingItem]) -> list[dict[str, Any]]:
        mapped: list[dict[str, Any]] = []
        failures: list[str] = []
        for item in items:
            record = None
            if item.internal_product_code:
                record = self._by_barcode.get(_normalize(item.internal_product_code))
                if record:
                    record = {**record, "matched_by": "explicit_internal_product_code"}
            if not record and item.platform_sku:
                record = self._by_sku.get(_normalize(item.platform_sku))
                if record:
                    record = {**record, "matched_by": "platform_sku"}
                else:
                    record = self._by_barcode.get(_normalize(item.platform_sku))
                    if record:
                        record = {**record, "matched_by": "product_barcode"}
            if not record:
                candidates = [item.variant_name, item.product_title]
                for candidate in candidates:
                    record = self._by_name.get(_normalize(candidate)) if candidate else None
                    if record:
                        record = {**record, "matched_by": "exact_product_name"}
                        break
            if not record:
                failures.append(item.platform_sku or item.variant_name or item.product_title)
                continue
            mapped.append(
                {
                    **record,
                    "fulfillment_sku": record["platform_sku"],
                    "quantity": item.quantity,
                    "declaration_name": item.declaration_name,
                    "declaration_value": item.declaration_value,
                    "declaration_quantity": item.declaration_quantity or item.quantity,
                    "source_title": item.product_title,
                }
            )

        if failures:
            raise SkuMappingError(
                "No deterministic SKU mapping for: " + ", ".join(failures)
            )
        return mapped

    @staticmethod
    def _locate_headers(rows: list[list[Any]]) -> tuple[int, dict[str, int | None]]:
        for row_index, row in enumerate(rows[:20]):
            headers = [str(value or "") for value in row]
            sku_index = _find_index(headers, PLATFORM_SKU_HEADERS)
            code_index = _find_index(headers, PRODUCT_CODE_HEADERS)
            name_index = _find_index(headers, PRODUCT_NAME_HEADERS)
            if code_index is not None and (sku_index is not None or name_index is not None):
                return row_index, {
                    "sku": sku_index,
                    "code": code_index,
                    "name": name_index,
                    "declaration_en": _find_index(headers, DECLARATION_EN_HEADERS),
                    "declaration_cn": _find_index(headers, DECLARATION_CN_HEADERS),
                    "declared_value": _find_index(headers, DECLARED_VALUE_HEADERS),
                    "weight": _find_index(headers, WEIGHT_HEADERS),
                }
        raise ConfigurationError("Could not find SKU/product headers in mapping file")

    @staticmethod
    def _record(row: list[Any], indexes: dict[str, int | None]) -> dict[str, str]:
        def value(key: str) -> str:
            index = indexes[key]
            return str(row[index] or "").strip() if index is not None and index < len(row) else ""

        return {
            "platform_sku": value("sku"),
            "product_code": value("code"),
            "product_name": value("name"),
            "declaration_en": value("declaration_en"),
            "declaration_cn": value("declaration_cn"),
            "declared_value": value("declared_value"),
            "weight": value("weight"),
        }
